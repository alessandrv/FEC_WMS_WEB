import pyodbc
from openpyxl import load_workbook
# Define your connection string
connection_string = 'DSN=fec;UID=informix;PWD=informix;'

def connect_to_db():
    """Connect to the database."""
    try:
        return pyodbc.connect(connection_string)
    except pyodbc.Error as e:
        raise Exception(f"Unable to connect to the database: {str(e)}")

def read_excel(file_path):
    """Read the Excel file and return a list of rows."""
    try:
        # Load the Excel workbook
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        dimensione = "Zero"

        # Read rows and skip the header
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the first row (header)
            codice_articolo, fornitore, movimento, locazione, quantita, _ = row  # Ignore PACCO_MULTIPLO
            print(row)
            if codice_articolo is not None:
                # Split the locazione into area, scaffale, colonna, and piano
                area, scaffale, colonna, piano = locazione.split('-')
                
                # Add the processed row to the list
                rows.append((codice_articolo, movimento, fornitore, area, scaffale, colonna, piano, quantita, "Zero"))
        print(rows)
        return rows
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")

def insert_data_to_db(connection, data):
    """Insert multiple rows into the wms_items table."""
    try:
        cursor = connection.cursor()

        # Define the SQL query for the insert
        insert_query = """
        INSERT INTO wms_items (id_art, id_mov, fornitore, area, scaffale, colonna, piano, qta, dimensione)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        # Execute the insert query with all rows
        cursor.executemany(insert_query, data)
        
        # Commit the transaction
        connection.commit()
        print("Data inserted successfully!")
    except Exception as e:
        connection.rollback()
        print(data)
        raise Exception(f"Error inserting data into database: {str(e)}")
    finally:
        cursor.close()

def main():
    # Path to the Excel file
    excel_file = input('Percorso file excel: ')

    # Connect to the database
    connection = connect_to_db()

    try:
        # Read data from the Excel file
        data = read_excel(excel_file)

        # Insert data into the database
        insert_data_to_db(connection, data)
    finally:
        connection.close()

if __name__ == '__main__':
    while True:
        main()
    
