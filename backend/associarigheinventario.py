import pandas as pd
from openpyxl import load_workbook

# Load the Excel files
file1_path = input("Inventario gestionale (.xlsm): ")
file2_path = input("Inventario eseguito (.xlsm): ")

# Read the first file using pandas
sheet1 = pd.read_excel(file1_path, sheet_name=0, engine='openpyxl')

# Read the second file using pandas
sheet2 = pd.read_excel(file2_path, sheet_name=0, engine='openpyxl')

# Normalize the column names for consistency
sheet1.columns = [col.strip() for col in sheet1.columns]
sheet2.columns = [col.strip() for col in sheet2.columns]

# Aggregate quantities from the second file
aggregated_quantities = (
    sheet2.groupby(sheet2.columns[0])[sheet2.columns[4]].sum().to_dict()
)

# Debug: Print aggregated quantities
print("Aggregated quantities from second file:")
for codice_articolo, total_quantita in aggregated_quantities.items():
    print(f"{codice_articolo}: {total_quantita}")

# Update the Quantità column (Column H) in the first file based on the aggregated quantities
updated_rows = []
for index, row in sheet1.iterrows():
    cod_art = str(row[sheet1.columns[1]]).strip().replace("=\"", "").replace("\"", "")
    if cod_art in aggregated_quantities:
        sheet1.iloc[index, 7] = aggregated_quantities[cod_art]  # Explicitly update Column H (zero-based index 7)
        updated_rows.append({
            'index': index,
            'cod_art': cod_art,
            'quantita': aggregated_quantities[cod_art]
        })

# Debug: Print updated rows
print("\nUpdated rows in the first file:")
for row_data in updated_rows:
    print(f"Row {row_data['index']}: Cod.Art {row_data['cod_art']} updated with Quantità {row_data['quantita']}")

# Load the .xlsm file to retain formatting and macros using openpyxl
wb = load_workbook(file1_path, keep_vba=True)

# Select the active sheet (or specify by name if needed)
ws = wb.active

# Update the cells in the Excel sheet based on the changes in pandas DataFrame
for row_data in updated_rows:
    # Adding 2 to account for header row and 0-based indexing
    ws.cell(row=row_data['index'] + 2, column=8, value=row_data['quantita'])

# Save the updated file (overwrite the existing file)
wb.save(file1_path)

print(f"\nUpdated file saved to {file1_path}")